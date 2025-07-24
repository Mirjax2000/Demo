"""View tests"""

import io
import shutil
import logging
import tempfile
from decimal import Decimal
from unittest.mock import patch
from openpyxl import load_workbook
from datetime import date, datetime

# --- django
from django.http import HttpRequest
from django.urls import reverse
from django.contrib.auth.models import User
from django.conf import settings
from django.contrib.messages import get_messages
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client as CL
from django.test import TestCase
from django.utils import timezone
from django.test import override_settings

# --- models
from app_sprava_montazi.models import (
    Article,
    Client,
    DistribHub,
    Order,
    OrderPDFStorage,
    Status,
    Team,
    TeamType,
    Upload,
)
from accounts.views import CustomLoginView
from app_sprava_montazi.views import OD_CHOICES, HUB_CHOICES


# --- utils
from ..utils import format_date, call_errors_adviced

# ---
test_logger = logging.getLogger("test_logger_login")
test_logger.setLevel(logging.INFO)
test_handler = (
    logging.StreamHandler()
)  # Or a FileHandler/NullHandler depending on your needs
test_logger.addHandler(test_handler)
test_logger.propagate = False  # Prevent logs from going to the root logger
LOGIN_URL = reverse("login")


class CustomLoginViewTests(TestCase):
    def setUp(self):
        # Create a test user for successful login attempts
        self.user = User.objects.create_user(
            username="testuser", password="testpassword123"
        )
        self.client = CL()

    def test_login_successful(self):
        """
        Test that a successful login logs an INFO message and redirects correctly.
        """
        # Patch your logger_login to capture its calls during the test
        with patch("accounts.views.logger_login", test_logger) as mock_logger:
            with self.assertLogs(mock_logger.name, level="INFO") as cm:
                response = self.client.post(
                    LOGIN_URL, {"username": "testuser", "password": "testpassword123"}
                )

                # Check if the user is authenticated
                self.assertTrue(response.wsgi_request.user.is_authenticated)
                self.assertRedirects(response, "/")  # Or your specific redirect URL

                # Verify the log message
                self.assertEqual(len(cm.output), 1)
                self.assertIn("login succeeded: testuser z IP:", cm.output[0])

    def test_login_failed_invalid_credentials(self):
        """
        Test that a failed login due to invalid credentials logs a WARNING message
        and stays on the login page.
        """
        with patch("accounts.views.logger_login", test_logger) as mock_logger:
            with self.assertLogs(mock_logger.name, level="WARNING") as cm:
                response = self.client.post(
                    LOGIN_URL, {"username": "testuser", "password": "wrongpassword"}
                )

                # Check that the user is not authenticated
                self.assertFalse(response.wsgi_request.user.is_authenticated)
                # Check that it re-renders the login page (status 200)
                self.assertEqual(response.status_code, 200)
                self.assertTemplateUsed(
                    response, "registration/login.html"
                )  # Or your login template

                # Verify the log message
                self.assertEqual(len(cm.output), 1)
                self.assertIn("login failed: testuser z IP:", cm.output[0])

    def test_login_failed_non_existent_user(self):
        """
        Test that a failed login due to a non-existent user logs a WARNING message.
        """
        with patch("accounts.views.logger_login", test_logger) as mock_logger:
            with self.assertLogs(mock_logger.name, level="WARNING") as cm:
                response = self.client.post(
                    LOGIN_URL,
                    {"username": "nonexistentuser", "password": "anypassword"},
                )

                self.assertFalse(response.wsgi_request.user.is_authenticated)
                self.assertEqual(response.status_code, 200)

                self.assertEqual(len(cm.output), 1)
                self.assertIn("login failed: nonexistentuser z IP:", cm.output[0])

    def test_get_client_ip_no_x_forwarded_for(self):
        """
        Test get_client_ip when HTTP_X_FORWARDED_FOR is not present.
        """

        request = self.client.get(LOGIN_URL).wsgi_request
        request.META["REMOTE_ADDR"] = "192.168.1.1"
        view = CustomLoginView()
        view.request = request
        self.assertEqual(view.get_client_ip(), "192.168.1.1")

    def test_get_client_ip_with_x_forwarded_for(self):
        """
        Test get_client_ip when HTTP_X_FORWARDED_FOR is present.
        """
        request = self.client.get(LOGIN_URL).wsgi_request
        request.META["HTTP_X_FORWARDED_FOR"] = "10.0.0.1, 192.168.1.5"
        request.META["REMOTE_ADDR"] = "192.168.1.1"  # This should be ignored
        view = CustomLoginView()
        view.request = request
        self.assertEqual(view.get_client_ip(), "10.0.0.1")


class IndexViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.url = reverse("index")
        self.template = "base.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")


class HomePageViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.url = reverse("homepage")
        self.template = "app_sprava_montazi/homepage/homepage.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "homepage")


class CreatePageViewTest(TestCase):
    """Nastavení testu: vytvoření uživatele a přihlášení."""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(username="testuser", password="testpass")

    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.url = reverse("createpage")
        self.template = "app_sprava_montazi/create/create.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_valid_form_submission(self):
        # Simulujeme upload CSV souboru
        csv_file = SimpleUploadedFile(
            "test.csv", b"col1,col2\nval1,val2", content_type="text/csv"
        )
        response = self.client.post(
            reverse("createpage"), {"file": csv_file}, follow=True
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Upload.objects.count(), 1)

    def test_invalid_form_submission(self):
        """bez zadani souboru do formulare"""

        response = self.client.post(reverse("createpage"), {}, follow=True)

        # Měla by se objevit chybová hláška
        self.assertContains(response, "Soubor je povinný!")
        self.assertEqual(Upload.objects.count(), 0)

    def test_upload_creates_db_entry(self):
        initial_count = Upload.objects.count()

        csv_file = SimpleUploadedFile(
            "valid.csv", b"header1,header2\n1,2", content_type="text/csv"
        )
        response = self.client.post(
            reverse("createpage"), {"file": csv_file}, follow=True
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Upload.objects.count(), initial_count + 1)

    @patch(
        "app_sprava_montazi.views.call_command", side_effect=KeyError("missing column")
    )
    @patch("app_sprava_montazi.views.DistribHub.objects.filter")
    def test_key_error_handling(self, mock_filter, mock_call_command):
        """
        Testuje, zda je KeyError při importu správně zachycen a zobrazena chybová zpráva.
        """
        mock_filter.return_value.exists.return_value = True

        csv_file = SimpleUploadedFile(
            "bad.csv", b"wrong,format\n1,2", content_type="text/csv"
        )

        response = self.client.post(self.url, {"file": csv_file}, follow=True)

        self.assertEqual(Upload.objects.count(), 1)
        upload = Upload.objects.get()
        expected_path = upload.file.path

        mock_call_command.assert_called_once_with("import_data", expected_path)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "Špatný soubor CSV")

    @patch(
        "app_sprava_montazi.views.call_command",
        side_effect=Exception("unexpected error"),
    )
    @patch("app_sprava_montazi.views.DistribHub.objects.filter")
    def test_unknown_exception_handling(self, mock_filter, mock_command):
        """
        Testuje, zda je obecná Exception při importu správně zachycena
        a zobrazena chybová zpráva "Neznamá chyba!".
        """
        mock_filter.return_value.exists.return_value = True

        csv_file = SimpleUploadedFile("bad.csv", b"ok", content_type="text/csv")

        response = self.client.post(self.url, {"file": csv_file}, follow=True)

        self.assertEqual(Upload.objects.count(), 1)
        upload = Upload.objects.get()
        expected_path = upload.file.path

        mock_command.assert_called_once_with("import_data", expected_path)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)  # Měla by být jen jedna zpráva
        self.assertIn("Neznamá chyba!", str(messages[0]))

    @patch(
        "app_sprava_montazi.views.call_command", side_effect=ValueError("wrong value")
    )
    @patch("app_sprava_montazi.views.DistribHub.objects.filter")
    def test_value_error_handling(self, mock_filter, mock_command):
        """
        Testuje, zda je ValueError při importu správně zachycen
        a zobrazena chybová zpráva.
        """

        mock_filter.return_value.exists.return_value = True
        csv_file = SimpleUploadedFile("bad.csv", b"ok", content_type="text/csv")

        response = self.client.post(self.url, {"file": csv_file}, follow=True)

        self.assertEqual(Upload.objects.count(), 1)
        upload = Upload.objects.get()
        expected_path = upload.file.path

        mock_command.assert_called_once_with("import_data", expected_path)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), "Chyba hodnoty v souboru CSV! wrong value")

    @patch("app_sprava_montazi.views.call_command", side_effect=DistribHub.DoesNotExist)
    @patch("app_sprava_montazi.views.DistribHub.objects.filter")
    def test_distribhub_doesnotexist_handling(self, mock_filter, mock_command):
        """
        Testuje, zda je DistribHub.DoesNotExist při importu správně zachycena
        a zobrazena chybová zpráva.
        """

        mock_filter.return_value.exists.return_value = True

        csv_file = SimpleUploadedFile("bad.csv", b"ok", content_type="text/csv")

        response = self.client.post(self.url, {"file": csv_file}, follow=True)

        self.assertEqual(Upload.objects.count(), 1)
        upload = Upload.objects.get()
        expected_path = upload.file.path

        mock_command.assert_called_once_with("import_data", expected_path)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

        messages = list(response.context["messages"])
        self.assertEqual(len(messages), 1)  #

        self.assertEqual(
            str(messages[0]).startswith("Chybné místo určení. Chyba:"), True
        )

    @patch("app_sprava_montazi.views.call_command")
    def test_distribhub_command_called_if_not_exist(self, mock_call_command):
        # Ujistíme se, že DistribHub je prázdný
        DistribHub.objects.all().delete()

        # Vytvoření falešného souboru pro formulář
        upload_file = SimpleUploadedFile(
            "test.csv", b"some,test,data\n1,2,3", content_type="text/csv"
        )
        response = self.client.post(self.url, data={"file": upload_file})

        # Ověření, že se command zavolal
        mock_call_command.assert_any_call("distrib_hub")
        mock_call_command.assert_any_call(
            "import_data", mock_call_command.call_args_list[1][0][1]
        )  # volání s file path

        # Úspěšné přesměrování
        self.assertRedirects(response, self.url)

        # Ověření zprávy
        messages = list(get_messages(response.wsgi_request))
        self.assertIn("Import dokončen.", str(messages[0]))

    @patch("app_sprava_montazi.views.call_command")
    def test_distribhub_command_not_called_if_exists(self, mock_call_command):
        # Vytvoříme záznam, takže DistribHub už existuje
        DistribHub.objects.create(code="652", city="Brno")

        upload_file = SimpleUploadedFile(
            "test.csv", b"some,test,data\n1,2,3", content_type="text/csv"
        )
        response = self.client.post(self.url, data={"file": upload_file})

        # Ověříme, že distrib_hub NEBYL zavolán
        command_names = [call[0][0] for call in mock_call_command.call_args_list]
        self.assertNotIn("distrib_hub", command_names)

        # Ale import_data ano
        self.assertIn("import_data", command_names)

        # Přesměrování
        self.assertRedirects(response, self.url)

        # Zpráva
        messages = list(get_messages(response.wsgi_request))
        self.assertIn(
            "Import dokončen. <strong>0</strong> nových zakázek",
            str(messages[0]),
        )


class DashboardViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.url = reverse("dashboard")
        self.template = "app_sprava_montazi/dashboard/dashboard.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "dashboard")


class ClientUpdateViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele

        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.template = "app_sprava_montazi/orders/order_update_client-form.html"
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="franta", zip_code="11111")
        self.order = Order.objects.create(
            order_number="703777143100437749-R",
            distrib_hub=self.hub,
            status=Status.NEW,
            client=self.customer,
            mandant="SCCZ",
            evidence_termin=date.today(),
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=None,
        )
        base_url = reverse("client_update", kwargs={"pk": self.order.client.pk})
        self.url = f"{base_url}?order_pk={self.order.pk}"

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_context_data(self):
        """test spravneho zobrazeni"""
        response = self.client.get(self.url)
        self.assertEqual(response.context["form_type"], "update")
        self.assertEqual(response.context["order"], self.order)
        self.assertQuerySetEqual(
            response.context["articles"],
            Article.objects.filter(order=self.order.pk),
            transform=lambda x: x,
        )
        self.assertEqual(response.context["active"], "orders_all")

    def test_success_message_displayed(self):
        data = {
            "name": "Karel",
            "zip_code": "12345",
        }
        response = self.client.post(self.url, data, follow=True)
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn("Zákazník: <strong>Karel</strong> aktualizován.", messages)

    def test_post_invalid_data_name(self):
        data = {
            "name": "",  # jméno je povinné
            "zip_code": "12345",
        }
        response = self.client.post(self.url, data)

        form = response.context["form"]
        self.assertEqual(response.status_code, 200)
        self.assertIn("Jméno je povinné!", str(form.errors["name"]))

    def test_post_invalid_data_zip_code(self):
        data = {
            "name": "ferda",  # jméno je povinné
            "zip_code": "",
        }
        response = self.client.post(self.url, data)

        form = response.context["form"]
        self.assertEqual(response.status_code, 200)
        self.assertIn("PSČ je povinné!", str(form.errors["zip_code"]))

    def test_post_valid_data(self):
        data = {
            "name": "Karel",
            "zip_code": "12345",
            "street": "Kopretinova 15",
            "city": "Kolin",
            "phone": "234234234",
            "email": "karel@seznam.cz",
        }
        response = self.client.post(self.url, data, follow=True)
        self.assertRedirects(
            response, reverse("order_detail", kwargs={"pk": self.order.pk})
        )
        self.customer.refresh_from_db()
        self.assertEqual(self.customer.name, "Karel")
        self.assertEqual(self.customer.zip_code, "12345")
        self.assertEqual(self.customer.street, "Kopretinova 15")
        self.assertEqual(self.customer.city, "Kolin")
        self.assertEqual(self.customer.phone, "234234234")
        self.assertEqual(self.customer.email, "karel@seznam.cz")


class OrderCreateViewTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.template = "app_sprava_montazi/orders/order_form.html"
        self.url = reverse("order_create")
        self.client = CL()
        self.client.login(username="testuser", password="testpass")
        self.valid_status_key = Status.NEW
        self.valid_team_type_key = TeamType.BY_ASSEMBLY_CREW
        self.hub = DistribHub.objects.create(code="111", city="Praha")
        self.customer = Client.objects.create(name="Pedro Pascal", zip_code="12345")
        self.team = Team.objects.create(
            name="Test Company",
            city="Praha",
            region="Střední Čechy",
            phone="123456789",
            email="test@company.cz",
            active=True,
            price_per_hour=150.50,
            price_per_km=12.30,
            notes="Toto je testovací poznámka.",
        )
        self.order = Order.objects.create(
            order_number="703777143100437749-R",
            distrib_hub=self.hub,
            status=Status.NEW,
            client=self.customer,
            mandant="SCCZ",
            evidence_termin=date.today(),
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=None,
        )

        # Základní validní data pro formuláře
        self.valid_customer_data = {
            "name": "Ferdinand",
            "street": "Kopretinova 15",
            "city": "Praha",
            "zip_code": "12345",
            "phone": "232232232",
            "email": "ferdinand@seznam.cz",
        }

        self.valid_order_data = {
            "order_number": "703777143100437749-O",
            "distrib_hub": self.hub.id,
            "mandant": "CCCC",
            "status": self.valid_status_key,
            "delivery_termin": "2024-05-15",
            "evidence_termin": "2024-05-20",
            "montage_termin": "2024-05-25T10:30",
            "team_type": self.valid_team_type_key,
            "team": self.team.id,
            "notes": "This is a test order note.",
        }

        self.formset_management_data = {
            "article_set-TOTAL_FORMS": 0,
            "article_set-INITIAL_FORMS": 0,
            "article_set-MIN_NUM_FORMS": 0,
            "article_set-MAX_NUM_FORMS": 1000,
        }

        self.valid_article_data = {
            "article_set-TOTAL_FORMS": 1,
            "article_set-INITIAL_FORMS": 0,
            "article_set-MIN_NUM_FORMS": 0,
            "article_set-MAX_NUM_FORMS": 1000,
            "article_set-0-name": "Postel",
            "article_set-0-quantity": 1,
            "article_set-0-price": 10000,
            "article_set-0-note": "Tvrda postel",
            "article_set-0-id": "",
            "article_set-0-DELETE": "",
        }

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá GET stránku
        a je použita správná šablona a formuláře v kontextu.
        """

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)
        # Zkontrolujeme, že formuláře jsou v kontextu
        self.assertIn("order_form", response.context)
        self.assertIn("client_form", response.context)
        self.assertIn("article_formset", response.context)
        # Zkontrolujeme, že formuláře v kontextu nejsou vázané na data (jsou prázdné)
        self.assertFalse(response.context["order_form"].is_bound)
        self.assertFalse(response.context["client_form"].is_bound)
        self.assertFalse(response.context["article_formset"].is_bound)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    @patch("django.contrib.messages.success")
    def test_successful_order_creation_with_articles(self, mock_success_message):
        """
        Testuje úspěšné vytvoření objednávky s klientem a články při validních datech.
        """
        initial_order_count = Order.objects.count()
        initial_client_count = Client.objects.count()
        initial_article_count = Article.objects.count()

        # Kombinujeme data pro POST - všechny formuláře validní, s jedním článkem
        post_data = {
            **self.valid_customer_data,
            **self.valid_order_data,
            **self.valid_article_data,
        }

        response = self.client.post(self.url, post_data)

        # Zkontrolujeme vazby
        created_order = Order.objects.latest("id")  # Získáme nově vytvořenou objednávku
        created_client = Client.objects.latest("id")  # Získáme nově vytvořeného klienta
        # Očekáváme přesměrování na detail objednávky
        self.assertEqual(response.status_code, 302)
        # Zkontrolujeme, že přesměrování je na URL detailu objednávky
        self.assertRedirects(
            response, reverse("order_detail", kwargs={"pk": created_order.pk})
        )

        # Zkontrolujeme, že data byla uložena do databáze
        self.assertEqual(Order.objects.count(), initial_order_count + 1)
        self.assertEqual(Client.objects.count(), initial_client_count + 1)
        self.assertEqual(
            Article.objects.count(),
            initial_article_count + self.valid_article_data["article_set-TOTAL_FORMS"],
        )

        self.assertEqual(created_order.client, created_client)
        self.assertEqual(
            created_order.articles.count(),
            self.valid_article_data["article_set-TOTAL_FORMS"],
        )

        created_article = created_order.articles.first()
        self.assertEqual(
            created_article.name, self.valid_article_data["article_set-0-name"]
        )
        self.assertEqual(
            created_article.quantity, self.valid_article_data["article_set-0-quantity"]
        )
        self.assertEqual(
            float(created_article.price), self.valid_article_data["article_set-0-price"]
        )  # Decimal vs float

        mock_success_message.assert_called_once()
        args, kwargs = mock_success_message.call_args
        msg: str = f"Objednávka vytvořena se zákazníkem: <strong>{created_order.client.name}</strong>."
        self.assertEqual(args[1], msg)

    @patch("django.contrib.messages.error")
    def test_invalid_order_data_shows_errors(self, mock_error_message):
        invalid_data = self.valid_order_data.copy()
        invalid_data["order_number"] = ""
        invalid_data["evidence_termin"] = ""
        invalid_data["distrib_hub"] = ""
        invalid_data["mandant"] = ""
        post_data = {
            **self.valid_customer_data,
            **invalid_data,
            **self.formset_management_data,
        }
        response = self.client.post(self.url, post_data)
        order_form = response.context.get("order_form")

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "číslo objednávky je povinné!", str(order_form.errors["order_number"])
        )
        self.assertIn("Mandant je povinný!", str(order_form.errors["mandant"]))
        self.assertIn(
            "Toto pole je vyžadováno", str(order_form.errors["evidence_termin"])
        )
        self.assertIn("místo určení je povinné!", str(order_form.errors["distrib_hub"]))

        mock_error_message.assert_called_once()
        args, kwargs = mock_error_message.call_args
        self.assertEqual(args[1], "Nastala chyba při ukládání objednávky.")

    @patch("django.contrib.messages.error")
    def test_unique_order_data_shows_errors(self, mock_error_message):
        invalid_data = self.valid_order_data.copy()
        invalid_data["order_number"] = "703777143100437749-R"
        post_data = {
            **self.valid_customer_data,
            **invalid_data,
            **self.formset_management_data,
        }
        response = self.client.post(self.url, post_data)
        order_form = response.context.get("order_form")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Objednávka už existuje!", str(order_form.errors["order_number"]))

        mock_error_message.assert_called_once()
        args, kwargs = mock_error_message.call_args
        self.assertEqual(args[1], "Nastala chyba při ukládání objednávky.")

    @patch("django.contrib.messages.success")
    def test_successful_order_creation_without_articles(self, mock_success_message):
        """
        Testuje úspěšné vytvoření objednávky s klientem, ale bez článků.
        Používá validní data vytvořená v setUp.
        """

        initial_order_count = Order.objects.count()
        initial_client_count = Client.objects.count()
        initial_article_count = Article.objects.count()

        # Kombinujeme data pro POST - validní Order a Client, prázdný formset
        # Zajistíme unikátní order_number pro tento test
        order_data = self.valid_order_data.copy()
        order_data["order_number"] = "703143100437749-R"
        post_data = {
            **self.valid_customer_data,
            **order_data,
            **self.formset_management_data,
        }

        response = self.client.post(self.url, post_data)
        created_order = Order.objects.latest("id")
        created_client = Client.objects.latest("id")

        self.assertEqual(response.status_code, 302)
        # Zkontrolujeme vazby

        self.assertEqual(response.status_code, 302)
        # Zkontrolujeme, že přesměrování je na URL detailu objednávky
        self.assertRedirects(
            response, reverse("order_detail", kwargs={"pk": created_order.pk})
        )

        self.assertEqual(Order.objects.count(), initial_order_count + 1)
        self.assertEqual(Client.objects.count(), initial_client_count + 1)
        self.assertEqual(Article.objects.count(), initial_article_count)

        self.assertEqual(created_order.client, created_client)
        self.assertEqual(created_order.articles.count(), 0)
        self.assertEqual(created_order.distrib_hub, self.hub)
        self.assertEqual(created_order.team, self.team)
        self.assertEqual(created_order.status, Status.ADVICED)
        self.assertEqual(created_order.team_type, self.valid_team_type_key)

        mock_success_message.assert_called_once()
        args, kwargs = mock_success_message.call_args
        msg: str = f"Objednávka vytvořena se zákazníkem: <strong>{created_order.client.name}</strong>."
        self.assertEqual(args[1], msg)

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")


class OrderDeleteViewTest(TestCase):
    def setUp(self) -> None:
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Pedro Pascal", zip_code="12345")
        self.team = Team.objects.create(
            name="Ferda Company",
            city="Praha",
            phone="234234234",
            email="ferda.company@gmail.cz",
        )
        self.order_hidden = Order.objects.create(
            order_number="703777143100437749-R",
            distrib_hub=self.hub,
            status=Status.HIDDEN,
            client=self.customer,
            mandant="SCCZ",
            evidence_termin=date.today(),
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=None,
        )
        self.order_new = Order.objects.create(
            order_number="703777143100437750-R",
            distrib_hub=self.hub,
            status=Status.NEW,
            client=self.customer,
            mandant="SCCZ",
            evidence_termin=date.today(),
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=None,
        )
        self.order_billed = Order.objects.create(
            order_number="703777143100437751-R",
            distrib_hub=self.hub,
            status=Status.BILLED,
            client=self.customer,
            mandant="SCCZ",
            evidence_termin=date.today(),
            delivery_termin=date.today(),
            montage_termin=timezone.make_aware(datetime(2025, 4, 10, 10, 0)),
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=self.team,
        )
        OrderPDFStorage.objects.create(
            order=self.order_billed, file=SimpleUploadedFile("file.pdf", b"dummy")
        )
        self.url_hidden = reverse("delete_order", kwargs={"pk": self.order_hidden.pk})
        self.url_new = reverse("delete_order", kwargs={"pk": self.order_new.pk})
        self.url_billed = reverse("delete_order", kwargs={"pk": self.order_billed.pk})

    def test_logged_in(self):
        response = self.client.post(self.url_hidden)
        self.assertRedirects(response, reverse("orders"))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Order.objects.filter(pk=self.order_hidden.pk).exists())

    def test_redirect_if_not_logged_in(self):
        self.client.logout()
        response = self.client.post(self.url_hidden)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url_hidden}")

    def test_success_message_displayed(self):
        response = self.client.post(self.url_hidden, follow=True)
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn(
            f"Zakázka: {self.order_hidden.order_number} byla smazána.", messages
        )

    def test_delete_order_with_pdf_storage_should_fail(self):
        # Navázat PDF na objednávku (ochrání ji před smazáním)
        OrderPDFStorage.objects.create(
            order=self.order_hidden, file=SimpleUploadedFile("file.pdf", b"dummy")
        )

        response = self.client.post(self.url_hidden, follow=True)

        self.assertTrue(Order.objects.filter(pk=self.order_hidden.pk).exists())

        # Ověřím, že se zobrazila chybová hláška
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn(
            f"Zakázku: {self.order_hidden.order_number} nelze smazat, protože má vazby na jiné záznamy.",
            messages,
        )
        # Cílová stránka po redirectu
        self.assertEqual(response.request["PATH_INFO"], reverse("orders"))

    def test_cant_delete_order_with_status_new(self):
        response = self.client.post(self.url_new)
        self.assertRedirects(response, reverse("orders"))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Order.objects.filter(pk=self.order_new.pk).exists())

    def test_cant_delete_order_with_status_new_with_message(self):
        response = self.client.post(self.url_new, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertRedirects(response, reverse("orders"))
        self.assertTrue(Order.objects.filter(pk=self.order_new.pk).exists())
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn("Jen skryté zakázky jde smazat!", messages)

    def test_cant_delete_order_with_status_billed(self):
        response = self.client.post(self.url_billed)
        self.assertRedirects(response, reverse("orders"))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Order.objects.filter(pk=self.order_billed.pk).exists())

    def test_cant_delete_order_with_status_billed_with_message(self):
        response = self.client.post(self.url_billed, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertRedirects(response, reverse("orders"))
        self.assertTrue(Order.objects.filter(pk=self.order_billed.pk).exists())
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn("Jen skryté zakázky jde smazat!", messages)


class OrdersAllViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.team = Team.objects.create(
            name="Ferda Company",
            city="Praha",
            phone="234234234",
            email="ferda.company@gmail.cz",
        )
        self.url = reverse("orders")
        self.template = "app_sprava_montazi/orders/orders_all.html"
        self.range: int = 10
        self.pagination: int = 15

        # Vytvoříme  objednávky
        # status new by assembly crew
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-1-{i}", zip_code=f"123{i:02}"
            )
            Order.objects.create(
                order_number=f"NEW_BY_CREW{i:05}-O",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 1, 1),
                delivery_termin=date(2025, 2, 10),
                status=Status.NEW,
                team_type=TeamType.BY_ASSEMBLY_CREW,
            )
        # status new by customer
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-2-{i}", zip_code=f"321{i:02}"
            )
            Order.objects.create(
                order_number=f"NEW_BY_CUST{i:05}-O",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 1, 1),
                delivery_termin=date(2025, 2, 10),
                status=Status.NEW,
                team_type=TeamType.BY_CUSTOMER,
            )
        # status adviced
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-3-{i}", zip_code=f"123{i:02}"
            )
            Order.objects.create(
                order_number=f"ADVICED-{i:06}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 2, 2),
                delivery_termin=date(2025, 3, 4),
                montage_termin=timezone.make_aware(datetime(2025, 4, 10, 10, 0)),
                status=Status.ADVICED,
                team_type=TeamType.BY_ASSEMBLY_CREW,
                team=self.team,
            )
        # status Hidden by customer
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-4-{i}", zip_code=f"321{i:02}"
            )
            Order.objects.create(
                order_number=f"HIDDEN_CUST{i:05}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 2, 2),
                delivery_termin=date(2025, 3, 4),
                status=Status.HIDDEN,
                team_type=TeamType.BY_CUSTOMER,
            )
        # status Realized
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-5-{i}", zip_code=f"123{i:02}"
            )
            Order.objects.create(
                order_number=f"REALIZED_{i:05}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 2, 2),
                delivery_termin=date(2025, 3, 4),
                montage_termin=timezone.make_aware(datetime(2025, 4, 10, 10, 0)),
                status=Status.REALIZED,
                team_type=TeamType.BY_ASSEMBLY_CREW,
                team=self.team,
            )
        # status Canceled
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-{i}", zip_code=f"321{i:02}"
            )
            Order.objects.create(
                order_number=f"CANCELED_{i:05}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 2, 2),
                delivery_termin=date(2025, 3, 4),
                montage_termin=timezone.make_aware(datetime(2025, 4, 10, 10, 0)),
                status=Status.CANCELED,
                team_type=TeamType.BY_ASSEMBLY_CREW,
                team=self.team,
            )
        # status Billed
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-6-{i}", zip_code=f"123{i:02}"
            )
            Order.objects.create(
                order_number=f"BILLED_{i:06}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2025, 2, 2),
                delivery_termin=date(2025, 3, 4),
                montage_termin=timezone.make_aware(datetime(2025, 4, 10, 10, 0)),
                status=Status.BILLED,
                team_type=TeamType.BY_ASSEMBLY_CREW,
                team=self.team,
            )
        # OD 701 NEW
        for i in range(self.range):
            customer = Client.objects.create(
                name=f"Customer test-7-{i}", zip_code=f"321{i:02}"
            )
            Order.objects.create(
                order_number=f"701{i:05}-R",
                distrib_hub=self.hub,
                mandant="SCCZ",
                client=customer,
                evidence_termin=date(2024, 2, 2),
                delivery_termin=date(2024, 3, 4),
                montage_termin=timezone.make_aware(datetime(2024, 4, 10, 10, 0)),
                status=Status.NEW,
                team_type=TeamType.BY_ASSEMBLY_CREW,
            )

    def test_filter_by_status_new(self):
        """filtruje status NEW"""
        response = self.client.get(self.url, {"status": "New"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_filter_by_status_adviced(self):
        """filtruje status Adviced"""
        response = self.client.get(self.url, {"status": "Adviced"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_filter_by_status_realized(self):
        """Filtruje status Realized"""
        response = self.client.get(self.url, {"status": "Realized"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_filter_by_status_od_701(self):
        """Filtruje OD 701"""
        response = self.client.get(self.url, {"od": "701"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_filter_by_start_date(self):
        response = self.client.get(self.url, {"start_date": "2024-01-10"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_filter_by_end_date(self):
        response = self.client.get(self.url, {"end_date": "2024-01-10"})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_combined_filters(self):
        response = self.client.get(
            self.url,
            {
                "status": "New",
                "od": "703",
                "start_date": "2024-01-01",
                "end_date": "2024-01-30",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_orders_all_view(self):
        """
        Test pro zobrazení všech objednávek.
        """

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)
        self.assertContains(response, 'name="invalid_filtr"', html=False)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        # invalid
        self.assertEqual(response.context["invalid"], False)
        # status
        self.assertEqual(response.context["statuses"], Status)
        self.assertEqual(response.context["raw_status"], "")
        self.assertEqual(response.context["get_status"], "")
        # mandant
        self.assertEqual(response.context["get_mandant"], None)
        # od
        self.assertEqual(response.context["od_choices"], OD_CHOICES)
        self.assertEqual(response.context["raw_od"], "")
        self.assertEqual(response.context["od_value"], "")
        # hub
        self.assertEqual(response.context["hub_choices"], HUB_CHOICES)
        self.assertEqual(response.context["raw_hub"], "")
        self.assertEqual(response.context["hub_value"], "")
        # start end
        self.assertEqual(response.context["get_start"], None)
        self.assertEqual(response.context["get_end"], None)

        self.assertEqual(response.context["request"].path, self.url)
        self.assertIsInstance(response.context["request"], HttpRequest)

    def test_active_context_variable_status_new(self):
        response = self.client.get(self.url, {"status": "New"})
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        # invalid
        self.assertEqual(response.context["invalid"], False)
        # status
        self.assertEqual(response.context["statuses"], Status)
        self.assertEqual(response.context["raw_status"], "New")
        self.assertEqual(response.context["get_status"], "Nový")
        # mandant
        self.assertEqual(response.context["get_mandant"], None)
        # od
        self.assertEqual(response.context["od_choices"], OD_CHOICES)
        self.assertEqual(response.context["raw_od"], "")
        self.assertEqual(response.context["od_value"], "")
        # start end
        self.assertEqual(response.context["get_start"], None)
        self.assertEqual(response.context["get_end"], None)
        # hub
        self.assertEqual(response.context["hub_choices"], HUB_CHOICES)
        self.assertEqual(response.context["raw_hub"], "")
        self.assertEqual(response.context["hub_value"], "")

        self.assertEqual(response.context["request"].path, self.url)
        self.assertIsInstance(response.context["request"], HttpRequest)

    def test_active_context_variable_status_new_manndant(self):
        response = self.client.get(self.url, {"status": "New", "mandant": "SCCZ"})
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        # invalid
        self.assertEqual(response.context["invalid"], False)
        # status
        self.assertEqual(response.context["statuses"], Status)
        self.assertEqual(response.context["raw_status"], "New")
        self.assertEqual(response.context["get_status"], "Nový")
        # mandant
        self.assertEqual(response.context["get_mandant"], "SCCZ")
        # od
        self.assertEqual(response.context["od_choices"], OD_CHOICES)
        self.assertEqual(response.context["raw_od"], "")
        self.assertEqual(response.context["od_value"], "")
        # hub
        self.assertEqual(response.context["hub_choices"], HUB_CHOICES)
        self.assertEqual(response.context["raw_hub"], "")
        self.assertEqual(response.context["hub_value"], "")
        # start end
        self.assertEqual(response.context["get_start"], None)
        self.assertEqual(response.context["get_end"], None)

        self.assertEqual(response.context["request"].path, self.url)
        self.assertIsInstance(response.context["request"], HttpRequest)

    def test_active_context_variable_status_new_manndant_701(self):
        response = self.client.get(
            self.url, {"status": "New", "mandant": "SCCZ", "od": "701"}
        )
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        # invalid
        self.assertEqual(response.context["invalid"], False)
        # status
        self.assertEqual(response.context["statuses"], Status)
        self.assertEqual(response.context["raw_status"], "New")
        self.assertEqual(response.context["get_status"], "Nový")
        # mandant
        self.assertEqual(response.context["get_mandant"], "SCCZ")
        # od
        self.assertEqual(response.context["od_choices"], OD_CHOICES)
        self.assertEqual(response.context["raw_od"], "701")
        self.assertEqual(response.context["od_value"], "OD Stodůlky")
        # hub
        self.assertEqual(response.context["hub_choices"], HUB_CHOICES)
        self.assertEqual(response.context["raw_hub"], "")
        self.assertEqual(response.context["hub_value"], "")
        # start end
        self.assertEqual(response.context["get_start"], None)
        self.assertEqual(response.context["get_end"], None)

        self.assertEqual(response.context["request"].path, self.url)
        self.assertIsInstance(response.context["request"], HttpRequest)

    def test_active_context_variable_distrib_hub_brno(self):
        response = self.client.get(self.url, {"status": "New", "hub": "626-chrastany"})
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        # invalid
        self.assertEqual(response.context["invalid"], False)
        # status
        self.assertEqual(response.context["statuses"], Status)
        self.assertEqual(response.context["raw_status"], "New")
        self.assertEqual(response.context["get_status"], "Nový")
        # mandant
        self.assertEqual(response.context["get_mandant"], None)
        # od
        self.assertEqual(response.context["od_choices"], OD_CHOICES)
        self.assertEqual(response.context["raw_od"], "")
        self.assertEqual(response.context["od_value"], "")
        # hub
        self.assertEqual(response.context["hub_choices"], HUB_CHOICES)
        self.assertEqual(response.context["raw_hub"], "626-chrastany")
        self.assertEqual(response.context["hub_value"], "626-Chrastany")
        # start end
        self.assertEqual(response.context["get_start"], None)
        self.assertEqual(response.context["get_end"], None)

        self.assertEqual(response.context["request"].path, self.url)
        self.assertIsInstance(response.context["request"], HttpRequest)

    # --- adviced nema odeslany email takze 10 spatne
    def test_invalid_filtr_is_true(self):
        response = self.client.get(self.url, {"invalid": "true"})
        self.assertEqual(response.status_code, 200)
        is_error, invalid_count = call_errors_adviced()
        self.assertTrue(is_error)
        self.assertEqual(invalid_count, 10)
        self.assertEqual(response.context["invalid"], True)


class TeamsViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.url = reverse("teams")
        self.template = "app_sprava_montazi/teams/teams_all.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "teams")


class TeamCreateTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.url = reverse("team_create")
        self.template = "app_sprava_montazi/teams/team_form.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "teams")


class TeamDetailViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.template = "app_sprava_montazi/teams/team_detail.html"
        self.active_team = Team.objects.create(
            name="active company",
            city="Praha",
            region="Střední Čechy",
            phone="123456789",
            email="active_company@company.cz",
            active=True,
            price_per_hour=150.50,
            price_per_km=12.30,
            notes="Toto je testovací poznámka.",
        )
        self.not_active_team = Team.objects.create(
            name="not active company",
            city="Praha",
            region="Střední Čechy",
            phone="123456789",
            email="not_active_company@company.cz",
            active=False,
            price_per_hour=150.50,
            price_per_km=12.30,
            notes="Toto je testovací poznámka.",
        )
        self.url_active = reverse("team_detail", kwargs={"slug": self.active_team.slug})
        self.url_not_active = reverse(
            "team_detail", kwargs={"slug": self.not_active_team.slug}
        )
        self.client.login(username="testuser", password="testpass")

    def test_logged_in_active(self):
        response = self.client.get(self.url_active)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in_active_team(self):
        self.client.logout()
        response = self.client.get(self.url_active)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url_active}")

    def test_redirect_if_not_logged_in_not_active_team(self):
        self.client.logout()
        response = self.client.get(self.url_not_active)
        self.assertRedirects(
            response, f"{settings.LOGIN_URL}?next={self.url_not_active}"
        )

    def test_active_context_variable_active(self):
        response = self.client.get(self.url_active)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "teams")

    def test_active_context_variable_not_active(self):
        response = self.client.get(self.url_not_active)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "teams")

    def test_content_in_html_not_active(self):
        response = self.client.get(self.url_not_active)
        self.assertIn("active", response.context)
        self.assertContains(response, 'id="checkBoxDeleteTeam"')
        self.assertContains(response, 'id="deleteTeamButton"')
        self.assertContains(response, 'name="delete-team"')

    def test_content_in_html_active(self):
        response = self.client.get(self.url_active)
        self.assertIn("active", response.context)
        self.assertNotContains(response, 'id="checkBoxDeleteTeam"')
        self.assertNotContains(response, 'id="deleteTeamButton"')
        self.assertNotContains(response, 'name="delete-team"')
        self.assertContains(response, 'name="no-delete-team"')


class TeamUpdateViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.team = Team.objects.create(
            name="Test Company",
            city="Praha",
            region="Střední Čechy",
            phone="234234234",
            email="test@company.cz",
            active=True,
            price_per_hour=Decimal(150.50),
            price_per_km=Decimal(12.30),
            notes="Toto je testovací poznámka.",
        )
        self.url = reverse("team_update", kwargs={"slug": self.team.slug})
        self.template = "app_sprava_montazi/teams/team_form.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "teams")

    def test_before_update(self):
        response = self.client.get(self.url)
        self.assertEqual(self.team.name, "Test Company")
        self.assertEqual(self.team.city, "Praha")
        self.assertEqual(self.team.region, "Střední Čechy")
        self.assertEqual(self.team.phone, "234234234")
        self.assertEqual(self.team.email, "test@company.cz")
        self.assertEqual(self.team.price_per_hour, Decimal(150.50))
        self.assertEqual(self.team.price_per_km, Decimal(12.30))
        self.assertEqual(self.team.notes, "Toto je testovací poznámka.")

    def test_team_update_post_1(self):
        """
        Testuje, že POST požadavek aktualizuje tým a přesměruje správně.
        """
        updated_data = {
            "city": "Brno",
            "region": "Jižní Morava",
            "phone": "234234234",
            "email": "updated@company.cz",
            "active": True,
            "price_per_hour": Decimal(200.00),
            "price_per_km": Decimal(15.00),
            "notes": "Aktualizovaná poznámka.",
        }

        response = self.client.post(self.url, updated_data, follow=True)
        # Dočasně si vypiš, jestli tam nebyla chyba
        self.team.refresh_from_db()

        # Ověříme, že data byla aktualizována
        self.assertEqual(self.team.name, "Test Company")
        self.assertEqual(self.team.city, "Brno")
        self.assertEqual(self.team.region, "Jižní Morava")
        self.assertEqual(self.team.phone, "234234234")
        self.assertEqual(self.team.email, "updated@company.cz")
        self.assertEqual(self.team.price_per_hour, 200.00)
        self.assertEqual(self.team.price_per_km, 15.00)
        self.assertEqual(self.team.notes, "Aktualizovaná poznámka.")

        # Ověříme, že došlo k přesměrování
        expected_url = reverse("team_detail", kwargs={"slug": self.team.slug})
        self.assertRedirects(response, expected_url)

        # Ověříme, že zpráva byla přidána
        messages_list = list(response.context["messages"])
        self.assertTrue(any("byl aktualizovan" in str(msg) for msg in messages_list))

    def test_team_update_post_2(self):
        updated_data = {
            "city": "Brno",
            "region": "Jižní Morava",
            "phone": "234234234",
            "email": "updated@company.cz",
            "active": False,
            "price_per_hour": Decimal(200.00),
            "price_per_km": Decimal(15.00),
            "notes": "Aktualizovaná poznámka.",
        }

        response = self.client.post(self.url, updated_data)

        # Záznam v DB zůstal nezměněný
        self.team.refresh_from_db()
        self.assertEqual(self.team.name, "Test Company")
        self.assertEqual(self.team.region, "Jižní Morava")
        self.assertEqual(self.team.phone, "234234234")
        self.assertEqual(self.team.email, "updated@company.cz")
        self.assertEqual(self.team.active, False)
        self.assertEqual(self.team.price_per_hour, Decimal(200.00))
        self.assertEqual(self.team.price_per_km, Decimal(15.00))
        self.assertEqual(self.team.notes, "Aktualizovaná poznámka.")


class ClientOrdersViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.customer = Client.objects.create(name="Franta Pina", zip_code="12345")
        self.url = reverse("client_orders", kwargs={"slug": self.customer.slug})
        self.template = "app_sprava_montazi/orders/client_orders.html"
        self.client.login(username="testuser", password="testpass")

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")


class OrderHistoryViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.template = "app_sprava_montazi/orders/order_detail_history.html"
        self.client.login(username="testuser", password="testpass")
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Franta test", zip_code="12345")
        self.order = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.NEW,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )
        self.url = reverse("order_history", kwargs={"pk": self.order.pk})

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")


class OrderDetailViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.template = "app_sprava_montazi/orders/order_detail.html"
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Franta test", zip_code="12345")
        self.order = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.NEW,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )
        self.url = reverse("order_detail", kwargs={"pk": self.order.pk})

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.template)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_active_context_variable(self):
        response = self.client.get(self.url)
        self.assertIn("active", response.context)
        self.assertEqual(response.context["active"], "orders_all")

    def test_non_existent_order_returns_404(self):
        url = reverse("order_detail", kwargs={"pk": 9999})  # Neexistující PK
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

    def test_order_in_context(self):
        response = self.client.get(self.url)
        self.assertEqual(response.context["order"], self.order)

    def test_order_number_rendered_in_template(self):
        response = self.client.get(self.url)
        self.assertContains(response, "12345-R")

    def test_articles_in_context(self):
        # Vytvoříme články (Articles) pro tuto objednávku
        Article.objects.create(order=self.order, name="Test Article 1")
        Article.objects.create(order=self.order, name="Test Article 2")

        response = self.client.get(self.url)

        # Zkontrolujeme, že v kontextu jsou 2 články
        self.assertEqual(len(response.context["articles"]), 2)
        self.assertEqual(response.context["articles"][0].name, "Test Article 1")
        self.assertEqual(response.context["articles"][1].name, "Test Article 2")

    def test_order_fields_in_context(self):
        response = self.client.get(self.url)
        order = response.context["order"]

        self.assertEqual(order.evidence_termin, date(2024, 1, 1))
        self.assertEqual(order.delivery_termin, date(2024, 4, 10))
        self.assertEqual(
            order.montage_termin, timezone.make_aware(datetime(2024, 5, 20, 8, 0))
        )
        self.assertEqual(order.status, Status.NEW)
        self.assertEqual(order.team_type, TeamType.BY_ASSEMBLY_CREW)

    def test_order_dates_rendered_in_template(self):
        response = self.client.get(self.url)
        self.assertContains(response, "01. 01. 2024")  # evidence_termin
        self.assertContains(response, "10. 04. 2024")  # delivery_termin
        self.assertContains(response, "20. 05. 2024 08:00")  # montage_termin

    def test_polozky_rendered(self):
        response = self.client.get(self.url)
        self.assertContains(response, "Nový")
        self.assertContains(response, "Montážníky")
        self.assertContains(response, str(self.customer))
        self.assertContains(response, "SCCZ")
        self.assertContains(response, str(self.hub))


class ExportOrdersExcelViewTest(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(
            name="Franta test",
            street="Testovací 1",
            city="TestMesto",
            zip_code="12345",
            phone="+420123456789",
            email="franta@example.com",
        )
        self.team = Team.objects.create(
            name="Montazni Spolecnost A",
            city="Praha",
            phone="+420987654321",
            price_per_hour=350.00,
            active=True,
        )

        # První objednávka s kompletními daty
        self.order1 = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.NEW,
            team_type=TeamType.BY_ASSEMBLY_CREW,
            team=self.team,
            notes="Nějaké poznámky k zakázce 1.",
        )
        self.order1.refresh_from_db()

        # Druhá objednávka s chybějícími daty pro testování prázdných buněk
        self.customer2 = Client.objects.create(
            name="Petra Nova",
            zip_code="67890",
            street="Testovací 2",
            city="Brno",
            phone="+420111222333",
            email="petra@example.com",
            incomplete=False,
        )
        self.order2 = Order.objects.create(
            order_number="67890-S",
            distrib_hub=self.hub,
            mandant="OtherMandant",
            client=self.customer2,
            evidence_termin=date(2023, 6, 15),
            delivery_termin=None,  # Test None values
            montage_termin=None,  # Test None values
            status=Status.REALIZED,
            team_type=TeamType.BY_DELIVERY_CREW,
            team=None,
            notes="",
        )
        self.order2.refresh_from_db()

        self.order3 = Order.objects.create(
            order_number="54321-S",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer2,
            evidence_termin=date(2024, 6, 15),
            delivery_termin=None,  # Test None values
            montage_termin=None,  # Test None values
            status=Status.NEW,
            team_type=TeamType.BY_DELIVERY_CREW,
            team=None,
            notes="",
        )
        self.order3.refresh_from_db()

        self.order_hidden = Order.objects.create(
            order_number="HIDDEN-Y",
            distrib_hub=self.hub,
            mandant="XYZ",
            client=self.customer,
            evidence_termin=date(2024, 2, 1),
            status=Status.HIDDEN,  # Status, který je defaultně filtrován
            team_type=TeamType.BY_CUSTOMER,
        )

        self.order_canceled = Order.objects.create(
            order_number="CANCELED-X",
            distrib_hub=self.hub,
            mandant="ABC",
            client=self.customer,
            evidence_termin=date(2024, 3, 1),
            status=Status.CANCELED,
            team_type=TeamType.BY_CUSTOMER,
        )

        self.order_adviced = Order.objects.create(
            order_number="ADVICED-X",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 3, 1),
            status=Status.ADVICED,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )

        self.url = reverse("order_export")

    def test_excel_file_content_all_visible_orders(self):
        """
        Testuje obsah vygenerovaného Excel souboru včetně hlaviček a dat pro všechny
        viditelné objednávky (které nejsou 'Hidden').
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        excel_file = io.BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active  # Aktivní list

        # Test Hlaviček (stejné jako předtím)
        expected_headers = [
            "Číslo zakázky",
            "Místo určení",
            "Mandant",
            "Stav",
            "Zákazník",
            "PSC",
            "Termín evidence",
            "Termín doručení",
            "Termín montáže",
            "Realizace kým",
            "Montážní tým",
            "Poznámky",
        ]
        actual_headers = [cell.value for cell in ws[1]]
        self.assertEqual(actual_headers, expected_headers)

        # Získání viditelných objednávek z DB a jejich seřazení pro ověření
        visible_orders = Order.objects.exclude(
            status__in=["Hidden", "Billed", "Canceled"]
        )

        # Test datových řádků

        current_order = visible_orders[0]
        row_values = [cell.value for cell in ws[2]]
        self.assertEqual(row_values[0], current_order.order_number)
        self.assertEqual(row_values[1], str(current_order.distrib_hub))
        self.assertEqual(row_values[2], current_order.mandant)
        self.assertEqual(row_values[3], current_order.get_status_display())
        self.assertEqual(row_values[4], str(current_order.client))
        self.assertEqual(row_values[5], current_order.client.zip_code)
        self.assertEqual(row_values[6], format_date(current_order.evidence_termin))
        # ZMĚNA: Pokud je delivery_termin None, očekáváme None z Excelu
        self.assertEqual(
            row_values[7],
            None
            if current_order.delivery_termin is None
            else format_date(current_order.delivery_termin),
        )
        # ZMĚNA: Pokud je montage_termin None, očekáváme None z Excelu
        self.assertEqual(
            row_values[8],
            None
            if current_order.montage_termin is None
            else format_date(current_order.montage_termin),
        )
        self.assertEqual(row_values[9], current_order.get_team_type_display())
        self.assertEqual(
            row_values[10], str(current_order.team) if current_order.team else None
        )  # Může být None
        self.assertEqual(
            row_values[11], current_order.notes if current_order.notes else None
        )  # "" read as None

        # Řádek 2 (prvni viditelná objednávka: self.order1)
        current_order = visible_orders[1]

        row_values = [cell.value for cell in ws[3]]
        self.assertEqual(row_values[0], current_order.order_number)
        self.assertEqual(row_values[1], str(current_order.distrib_hub))
        self.assertEqual(row_values[2], current_order.mandant)
        self.assertEqual(row_values[3], current_order.get_status_display())
        self.assertEqual(row_values[4], str(current_order.client))
        self.assertEqual(row_values[5], current_order.client.zip_code)
        self.assertEqual(row_values[6], format_date(current_order.evidence_termin))
        # ZMĚNA
        self.assertEqual(
            row_values[7],
            None
            if current_order.delivery_termin is None
            else format_date(current_order.delivery_termin),
        )
        # ZMĚNA
        self.assertEqual(
            row_values[8],
            None
            if current_order.montage_termin is None
            else format_date(current_order.montage_termin),
        )
        self.assertEqual(row_values[9], current_order.get_team_type_display())
        self.assertEqual(
            row_values[10], str(current_order.team) if current_order.team else None
        )  # Může být None
        self.assertEqual(
            row_values[11], current_order.notes if current_order.notes else None
        )  # "" read as None

        current_order = visible_orders[2]
        row_values = [cell.value for cell in ws[4]]
        self.assertEqual(row_values[0], current_order.order_number)
        self.assertEqual(row_values[1], str(current_order.distrib_hub))
        self.assertEqual(row_values[2], current_order.mandant)
        self.assertEqual(row_values[3], current_order.get_status_display())
        self.assertEqual(row_values[4], str(current_order.client))
        self.assertEqual(row_values[5], current_order.client.zip_code)
        self.assertEqual(row_values[6], format_date(current_order.evidence_termin))
        self.assertEqual(row_values[7], format_date(current_order.delivery_termin))
        self.assertEqual(row_values[8], format_date(current_order.montage_termin))
        self.assertEqual(row_values[9], current_order.get_team_type_display())
        self.assertEqual(
            row_values[10], str(current_order.team) if current_order.team else None
        )
        self.assertEqual(
            row_values[11], current_order.notes if current_order.notes else None
        )

        # Zajistíme správný počet řádků (1 hlavička + 4 datové řádky)
        self.assertEqual(ws.max_row, 5)

    def test_excel_file_no_matching_orders(self):
        """
        Testuje generování Excelu, když žádná objednávka neodpovídá filtru.
        Měl by obsahovat pouze hlavičky.
        """
        response = self.client.get(self.url, {"status": Status.NEW, "od": "NO-MATCH"})
        self.assertEqual(response.status_code, 200)

        excel_file = io.BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        self.assertEqual(ws.max_row, 1)  # Only headers
        expected_headers = [
            "Číslo zakázky",
            "Místo určení",
            "Mandant",
            "Stav",
            "Zákazník",
            "PSC",
            "Termín evidence",
            "Termín doručení",
            "Termín montáže",
            "Realizace kým",
            "Montážní tým",
            "Poznámky",
        ]
        actual_headers = [cell.value for cell in ws[1]]
        self.assertEqual(actual_headers, expected_headers)
        # ZDE JE ZMĚNA: Očekávaný suffix je nyní "New_NO-MATCH_None_None"
        disposition = response["Content-Disposition"]
        self.assertIn(f"filename=objednavky-{Status.NEW.value}_", disposition)
        self.assertIn("NO-MATCH", disposition)
        self.assertIn(".xlsx", disposition)

    def test_excel_file_with_status_filter(self):
        """
        Testuje generování Excelu s použitím filtru 'status'.
        Měl by obsahovat pouze objednávky s daným statusem (zde ADVICED).
        """
        response = self.client.get(self.url, {"status": Status.ADVICED})
        self.assertEqual(response.status_code, 200)

        excel_file = io.BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        self.assertEqual(ws.max_row, 3)

        # Zkontrolujeme, zda se jedná o správnou objednávku
        order_row_values = [cell.value for cell in ws[2]]
        self.assertEqual(order_row_values[0], self.order_adviced.order_number)
        self.assertEqual(
            order_row_values[3], self.order_adviced.get_status_display()
        )  # ADVICED
        self.assertEqual(
            order_row_values[6], format_date(self.order_adviced.evidence_termin)
        )
        disposition = response["Content-Disposition"]
        self.assertIn(f"filename=objednavky-{Status.ADVICED.value}_", disposition)
        self.assertIn(".xlsx", disposition)


class PdfViewTestV1(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.url = reverse("mandant_pdf", kwargs={"mandant": "SCCZ"})

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_pdf_view_returns_pdf(self):
        url = reverse("mandant_pdf", kwargs={"mandant": "SCCZ"})
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("filename=", response["Content-Disposition"])
        self.assertGreater(len(response.content), 100)  # nějaký obsah
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertIn("Protokol_SCCZ.pdf", response["Content-Disposition"])


class PdfViewTestV2(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.url = reverse("mandant_pdf", kwargs={"mandant": "default"})

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")

    def test_pdf_view_returns_pdf(self):
        url = reverse("mandant_pdf", kwargs={"mandant": "default"})
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("filename=", response["Content-Disposition"])
        self.assertGreater(len(response.content), 100)  # nějaký obsah
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertIn("Protokol_default.pdf", response["Content-Disposition"])


class OrderPdfViewTestV1(TestCase):
    def setUp(self):
        # Vytvoříme testovacího uživatele
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")
        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Franta test", zip_code="12345")
        self.order = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.ADVICED,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )
        self.url = reverse("protocol_pdf", kwargs={"pk": self.order.pk})

    def test_logged_in(self):
        """
        Testuje, zda přihlášený uživatel úspěšně získá indexovou stránku
        a je použita správná šablona.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

    def test_redirect_if_not_logged_in(self):
        """
        Testuje, zda je uživatel přesměrován na přihlašovací stránku,
        pokud není přihlášen a pokusí se zobrazit indexovou stránku.
        """
        self.client.logout()
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url}")


class GeneratePDFViewTest(TestCase):
    def setUp(self):
        self._temp_media = tempfile.mkdtemp()  # vytvoří dočasný adresář
        self.override = override_settings(MEDIA_ROOT=self._temp_media)
        self.override.enable()
        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")

        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Franta test", zip_code="12345")
        self.team_active = Team.objects.create(
            name="Active Company",
            city="Praha",
            region="Střední Čechy",
            phone="123456789",
            email="active@company.cz",
            active=True,
            price_per_hour=150.50,
            price_per_km=12.30,
            notes="Toto je testovací poznámka.",
        )
        self.team_not_active = Team.objects.create(
            name="Not active company",
            city="Praha",
            region="Střední Čechy",
            phone="123456789",
            email="not_active@company.cz",
            active=False,
            price_per_hour=150.50,
            price_per_km=12.30,
            notes="Toto je testovací poznámka.",
        )
        self.order_with_active_team = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.ADVICED,
            team=self.team_active,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )
        self.order_with_not_active_team = Order.objects.create(
            order_number="54321-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.ADVICED,
            team=self.team_not_active,
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )
        self.url_active = reverse(
            "generate_pdf", kwargs={"pk": self.order_with_active_team.pk}
        )
        self.url_not_active = reverse(
            "generate_pdf", kwargs={"pk": self.order_with_not_active_team.pk}
        )

    def test_logged_in_redirects(self):
        """
        Přihlášený uživatel je přesměrován po úspěšném vygenerování PDF.
        """
        response = self.client.get(self.url_active)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(
            response, reverse("protocol", kwargs={"pk": self.order_with_active_team.pk})
        )

    def test_redirect_if_not_logged_in(self):
        """
        Nepřihlášený uživatel je přesměrován na login stránku.
        """
        self.client.logout()
        response = self.client.get(self.url_active)
        self.assertRedirects(response, f"{settings.LOGIN_URL}?next={self.url_active}")

    def test_pdf_is_generated_and_saved_v1(self):
        response = self.client.get(self.url_active)
        self.assertRedirects(
            response, reverse("protocol", kwargs={"pk": self.order_with_active_team.pk})
        )

        pdf_entry = OrderPDFStorage.objects.get(order=self.order_with_active_team)
        self.assertIsNotNone(pdf_entry.file)
        self.assertTrue(pdf_entry.file.name.endswith(".pdf"))
        self.assertTrue(default_storage.exists(pdf_entry.file.name))

        with default_storage.open(pdf_entry.file.name, "rb") as f:
            content = f.read()
            self.assertGreater(len(content), 0)

    def test_pdf_is_generated_with_message(self):
        response = self.client.get(self.url_active, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertRedirects(
            response, reverse("protocol", kwargs={"pk": self.order_with_active_team.pk})
        )
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn(
            (
                f"PDF: <strong>{str(self.order_with_active_team).upper()}</strong> "
                f"byl úspěšně vygenerován a uložen."
            ),
            messages,
        )

    def test_pdf_is_generated_with_message_not_active(self):
        response = self.client.get(self.url_not_active, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertRedirects(
            response,
            reverse("protocol", kwargs={"pk": self.order_with_not_active_team.pk}),
        )
        messages = [msg.message for msg in get_messages(response.wsgi_request)]
        self.assertIn("Nelze generovat protokol, Tým je neaktivní", messages)

    def tearDown(self):
        self.override.disable()

        # Smaž všechny soubory v dočasném MEDIA_ROOT
        shutil.rmtree(self._temp_media, ignore_errors=True)


class CheckPDFProtocolViewTest(TestCase):
    def setUp(self):
        # Vytvoření dočasného MEDIA_ROOT
        self._temp_media = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self._temp_media)
        self.override.enable()

        self.user = User.objects.create_user(username="testuser", password="testpass")
        self.client.login(username="testuser", password="testpass")

        self.hub = DistribHub.objects.create(code="626", city="Chrastany")
        self.customer = Client.objects.create(name="Franta test", zip_code="12345")

        self.order = Order.objects.create(
            order_number="12345-R",
            distrib_hub=self.hub,
            mandant="SCCZ",
            client=self.customer,
            evidence_termin=date(2024, 1, 1),
            delivery_termin=date(2024, 4, 10),
            montage_termin=timezone.make_aware(datetime(2024, 5, 20, 8, 0)),
            status=Status.ADVICED,
            team=Team.objects.create(
                name="Test Company",
                city="Praha",
                region="Střední Čechy",
                phone="123456789",
                email="test@company.cz",
                active=True,
                price_per_hour=150.50,
                price_per_km=12.30,
                notes="Toto je testovací poznámka.",
            ),
            team_type=TeamType.BY_ASSEMBLY_CREW,
        )

        self.test_pdf_content = b"%PDF-1.4 fake content for testing"
        self.test_pdf_file = SimpleUploadedFile(
            "test.pdf", self.test_pdf_content, content_type="application/pdf"
        )

        self.pdf_storage = OrderPDFStorage.objects.create(
            order=self.order,
            file=self.test_pdf_file,
        )

        self.url = reverse("check_pdf", kwargs={"pk": self.order.pk})

    def tearDown(self):
        # Úklid: odstraní vytvořený dočasný adresář a vypne override
        shutil.rmtree(self._temp_media)
        self.override.disable()

    def test_check_pdf_protocol_view(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        content = b"".join(response.streaming_content)
        self.assertIn(b"%PDF-1.4", content)
